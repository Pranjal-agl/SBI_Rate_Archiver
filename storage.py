"""
Storage layer for forex rate data.

Handles:
  - Daily Excel file:  data/<year>/forex_rates_YYYY-MM-DD.xlsx
  - Master CSV:        all_rates_history.csv  (append-only, deduped by date+bank+label)
"""

import logging
import os
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models import RateRecord

logger = logging.getLogger(__name__)

MASTER_CSV = Path("all_rates_history.csv")
DATA_ROOT = Path("data")


# ── helpers ───────────────────────────────────────────────────────────────────

def _daily_excel_path(for_date: date) -> Path:
    year_dir = DATA_ROOT / str(for_date.year)
    year_dir.mkdir(parents=True, exist_ok=True)
    return year_dir / f"forex_rates_{for_date.isoformat()}.xlsx"


def _style_workbook(wb: Workbook) -> None:
    """Apply professional formatting to the active worksheet."""
    ws = wb.active
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for cell in row:
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4


# ── public API ────────────────────────────────────────────────────────────────

def save_daily_excel(records: list[RateRecord], for_date: date | None = None) -> Path:
    """
    Write records to a date-stamped Excel file.

    Args:
        records:  List of RateRecord objects to persist.
        for_date: Override the target date (defaults to today).

    Returns:
        Path to the written file.
    """
    target_date = for_date or date.today()
    path = _daily_excel_path(target_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Forex Rates"
    ws.append(RateRecord.csv_header())

    for rec in records:
        ws.append(rec.as_row())

    _style_workbook(wb)
    wb.save(path)
    logger.info("Daily Excel saved → %s  (%d rows)", path, len(records))
    return path


def append_to_master_csv(records: list[RateRecord]) -> None:
    """
    Append records to the master historical CSV, deduplicating on
    (Date, Bank, Slab_Type) to prevent double-writes on re-runs.
    """
    new_df = pd.DataFrame([r.as_row() for r in records], columns=RateRecord.csv_header())

    if MASTER_CSV.exists():
        existing = pd.read_csv(MASTER_CSV, dtype=str)
        combined = pd.concat([existing, new_df.astype(str)], ignore_index=True)
        combined.drop_duplicates(subset=["Date", "Bank", "Slab_Type"], keep="last", inplace=True)
        combined.to_csv(MASTER_CSV, index=False)
        added = len(combined) - len(existing)
        logger.info("Master CSV updated → %s  (+%d new rows, %d total)", MASTER_CSV, added, len(combined))
    else:
        new_df.to_csv(MASTER_CSV, index=False)
        logger.info("Master CSV created → %s  (%d rows)", MASTER_CSV, len(new_df))


def load_history() -> pd.DataFrame:
    """Load the master CSV as a typed DataFrame. Returns empty DF if file missing."""
    if not MASTER_CSV.exists():
        return pd.DataFrame(columns=RateRecord.csv_header())
    df = pd.read_csv(MASTER_CSV)
    df["Date"] = pd.to_datetime(df["Date"])
    df["Rate"] = pd.to_numeric(df["Rate"], errors="coerce")
    return df.dropna(subset=["Rate"]).sort_values("Date")
